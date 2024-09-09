import qrcode
import openpyxl

#Gerar QR Code
qr = qrcode.QRCode(
    version=1,
    error_correction=qrcode.constants.ERROR_CORRECT_L,
    box_size=5,
    border=1,
)
qr.make(fit=False)

#Puxar Dados Excel
path = "C:\\Users\\User\\Downloads\\Qr code\\Dados.xlsx" #Altere para o Caminho da Planilha com os dados.
planilha = openpyxl.load_workbook(path)
celula = planilha.active

#Gerar QR Code com Dados do Excel
for i in range(4): #Coloque a quantidade de QR's a serem gerados.
    item_1 = celula.cell(row = i+1, column = 1) #Altere indicando a celula onde os dados serão coletados.
    item_2 = celula.cell(row = i+1, column = 3)
    texto_do_qr_code = f'''
    Qr Code de Exemplo
    COLUNA 1: {item_1.value}
    COLUNA 3: {item_2.value}
    '''
    qr.add_data(texto_do_qr_code)
    img = qr.make_image(fill_color="black", back_color="transparent")
    img.save(f"C:\\Users\\User\\Downloads\\Qr code\\QR's CODE\\QR.{str((i+1)).zfill(4)}.png") #Altere para o Caminho onde os QR's Codes devem ser salvos.
    qr.clear()